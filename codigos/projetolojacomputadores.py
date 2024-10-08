# -*- coding: utf-8 -*-
"""projetoLojaComputadores.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1_kAQqYmyG5TABt4bQoqU-0SH5apqRvNY
"""

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

# Importando os dados CSV combinados
combined_df = pd.read_csv('loja_computadores.csv')

# Exibindo os nomes das colunas para verificação
print("Colunas disponíveis no DataFrame:", combined_df.columns)

# Exibindo as primeiras linhas dos dados
print(combined_df.head())

# Tratamento de dados: convertendo preços para centavos
combined_df['Price_cents'] = combined_df['Price'] * 100  #Converte o preço de reais para centavos e armazena na nova coluna 'Price_cents'.

# Média de preços de todos os produtos
average_price = combined_df['Price'].mean()
print(f"Média dos preços de todos os produtos: ${average_price:.2f}")

# Média de preços por fabricante
manufacturer_avg_price = combined_df.groupby('Manufacturer_Name')['Price'].mean()
print(f"Média de preços por fabricante:\n{manufacturer_avg_price}")

# Produtos com preço superior a $180
products_above_180 = combined_df[combined_df['Price'] >= 180]
print(f"Produtos com preço >= $180:\n{products_above_180}")

# Exibindo todas as linhas dos dados (se o DataFrame não for muito grande)
print("Exibindo todas as linhas do DataFrame:")
print(combined_df)

# Exemplo de visualizações

# 1. Distribuição dos preços dos produtos
plt.figure(figsize=(10, 6)) # Define o tamanho da área do gráfico.
sns.histplot(combined_df['Price_cents'], bins=20, kde=True) # Plota um histograma dos preços em centavos com uma curva de densidade.
plt.title('Distribuição dos Preços dos Produtos') #  Define o título do gráfico.
plt.xlabel('Preço (em centavos)')# Define o rótulo do eixo X.
plt.ylabel('Frequência') # Define o rótulo do eixo Y.
# Ajusta o layout do gráfico para evitar sobreposição de elementos
plt.tight_layout()

# Salva o gráfico como uma imagem PNG no diretório /content
plt.savefig('distribuicao_precos.png')

# Fecha o gráfico atual para liberar memória e evitar problemas com gráficos subsequentes
plt.show() # Exibe o gráfico gerado.
plt.close()

# 2. Preço médio dos produtos por fabricante
plt.figure(figsize=(12, 8))  # Define o tamanho da figura do gráfico (largura de 12 polegadas e altura de 8 polegadas)

# Calcula o preço médio dos produtos para cada fabricante, agrupando por 'Manufacturer_Name'
# Ordena os resultados em ordem crescente de preço médio
avg_price_per_manufacturer = combined_df.groupby('Manufacturer_Name')['Price'].mean().sort_values() # Plota um gráfico de barras com os preços médios por fabricante
sns.barplot(x=avg_price_per_manufacturer.index,  # Define os rótulos do eixo X como os nomes dos fabricantes
            y=avg_price_per_manufacturer.values,  # Define os valores do eixo Y como os preços médios
            hue=avg_price_per_manufacturer.values,
            palette=sns.color_palette('viridis', n_colors=4))  # Define a paleta de cores para o gráfico de barras

plt.title('Preço Médio dos Produtos por Fabricante')  # Define o título do gráfico
plt.xlabel('Fabricante')  # Define o rótulo do eixo X
plt.ylabel('Preço Médio')  # Define o rótulo do eixo Y
plt.xticks(rotation=45)  # Rotaciona os rótulos do eixo X em 45 graus para melhor legibilidade
# Ajusta o layout do gráfico para evitar sobreposição de elementos
plt.tight_layout()
plt.savefig('preco_medio_fabricante.png')  # Salva o gráfico como uma imagem PNG
plt.show()  # Exibe o gráfico gerado
plt.close()  # Fecha o gráfico atual para liberar memória

# Define o tamanho da figura do gráfico (largura de 12 polegadas e altura de 8 polegadas)
plt.figure(figsize=(12, 8))

# Conta o número de produtos por fabricante usando a coluna 'Manufacturer_Name' e retorna uma série com fabricantes como índices e contagens como valores
product_count_per_manufacturer = combined_df['Manufacturer_Name'].value_counts() # Plota um gráfico de barras com o número de produtos por fabricante
sns.barplot(x=product_count_per_manufacturer.index, # x: Define os rótulos do eixo X como os fabricantes
            y=product_count_per_manufacturer.values, # y: Define os valores do eixo Y como o número de produtos por fabricante
            hue=product_count_per_manufacturer.values, # hue: Adiciona uma dimensão adicional para a coloração das barras; neste caso, usa os valores da contagem dos produtos, mas não afeta a coloração
            palette=sns.color_palette('plasma', n_colors=3)) # palette: Define a paleta de cores para o gráfico de barras usando a paleta 'plasma' e especifica que a paleta deve ter 3 cores

# Define o título do gráfico
plt.title('Contagem de Produtos por Fabricante')

# Define o rótulo do eixo X
plt.xlabel('Fabricante')

# Define o rótulo do eixo Y
plt.ylabel('Número de Produtos')

# Rotaciona os rótulos do eixo X em 45 graus para melhor legibilidade
plt.xticks(rotation=45)
# Ajusta o layout do gráfico para evitar sobreposição de elementos
plt.tight_layout()
# Salva o gráfico como uma imagem PNG
plt.savefig('contagem_produtos_fabricante.png')
# Exibe o gráfico gerado
plt.show()
# Fecha o gráfico atual para liberar memória
plt.close()

print(combined_df.shape)  # Mostra o número de linhas e colunas

# Salvando o DataFrame combinado em um arquivo CSV
combined_df.to_csv('dados_combinados.csv', index=False)

# Salvando o DataFrame combinado em um arquivo Excel
combined_df.to_excel('dados_combinados.xlsx', index=False)

# Salvando os dados em um arquivo Excel
with pd.ExcelWriter('dados_completos.xlsx') as writer:
    combined_df.to_excel(writer, sheet_name='Dados', index=False)

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Caminho das imagens
imagens = [
    'distribuicao_precos.png',
    'preco_medio_fabricante.png',  # Adicione mais caminhos de imagens
    'contagem_produtos_fabricante.png',

]

# Carregar o arquivo Excel existente
workbook = load_workbook('dados_completos.xlsx')

# Adicionar uma nova aba para gráficos
worksheet = workbook.create_sheet(title='Gráficos')

# Adicionar imagens verticalmente em uma única coluna
for i, imagem in enumerate(imagens):
    img = Image(imagem)
    cell = f'A{i * 30 + 1}'  # Ajuste o espaçamento vertical conforme necessário
    worksheet.add_image(img, cell)

# Salvar o arquivo Excel com as imagens
workbook.save('dados_completos_com_graficos.xlsx')  #tem que se aperfecôar porque fica sem medida exata nos graficos

"""exemplo para gerar codigos random en caso de precisar aumentar la db, deve ser trabalhado ainda"""

# # Gerar dados aleatórios
# num_new_entries = 10  # Número de novas entradas que você deseja adicionar

# # Gerar IDs únicos para os novos produtos
# new_ids = np.arange(combined_df['Product_Code'].max() + 1, combined_df['Product_Code'].max() + num_new_entries + 1)

# # Gerar nomes aleatórios
# new_names = [f'Produto_{i}' for i in range(num_new_entries)]

# # Gerar preços aleatórios entre 50 e 500
# new_prices = np.random.uniform(50, 500, size=num_new_entries)

# # Gerar IDs de fabricantes aleatórios (considerando que temos 6 fabricantes)
# new_manufacturers = np.random.randint(1, 7, size=num_new_entries)

# # Criar um novo DataFrame com dados aleatórios
# new_products_df = pd.DataFrame({
#     'Product_Code': new_ids,
#     'Name': new_names,
#     'Price': new_prices,
#     'Manufacturer': new_manufacturers
# })

# # Concatenar o DataFrame existente com o novo DataFrame
# combined_df = pd.concat([combined_df, new_products_df], ignore_index=True)

